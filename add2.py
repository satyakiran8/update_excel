import openpyxl
import os

def update_excel(name, age, city, phone_number, file_path="data.xlsx"):
    """
    Add a new row to Excel file with automatic s_no generation.
    Only appends new row - doesn't rewrite existing data.
    
    Args:
        name (str): Person's name
        age (int): Person's age  
        city (str): Person's city
        phone_number (int): Person's phone number
        file_path (str): Path to Excel file (default: "data.xlsx")
    """
    
    # Input validation
    try:
        # Validate name
        if not isinstance(name, str) or not name.strip():
            
            return "❌ Error: Name must be a non-empty string"
            
        # Validate age
        age = int(age)
        if age <= 0:
            print("❌ Error: Age must be a positive number")
            return "❌ Error: Age must be a positive number"
            
        # Validate city
        if not isinstance(city, str) or not city.strip():
            
            return "❌ Error: City must be a non-empty string"
            
        # Validate phone_number
        phone_number = int(phone_number)
        
    except ValueError:
        
        return "❌ Error: Age and Phone_number must be valid numbers"
    
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            # Create new workbook with headers
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Add headers
            headers = ["s_no", "Name", "Age", "City", "Phone_number"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                
            print(f"📄 Created new Excel file: {file_path}")
            next_s_no = 1
        else:
            # Load existing workbook
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Find the next s_no automatically
            max_row = ws.max_row
            if max_row > 1:  # If there's data beyond headers
                last_s_no = ws.cell(row=max_row, column=1).value
                next_s_no = last_s_no + 1 if last_s_no else 1
            else:
                next_s_no = 1
        
        # Find the next empty row
        next_row = ws.max_row + 1
        
        # Add ONLY the new row data (no rewriting of old data)
        ws.cell(row=next_row, column=1, value=next_s_no)
        ws.cell(row=next_row, column=2, value=name.strip())
        ws.cell(row=next_row, column=3, value=age)
        ws.cell(row=next_row, column=4, value=city.strip())
        ws.cell(row=next_row, column=5, value=phone_number)
        
        # Save the file
        wb.save(file_path)
        
        print(f"✅ SUCCESS! Appended row {next_s_no}: {name} | {age} | {city} | {phone_number}")
        return True
        
    except PermissionError:
        
        return "❌ Error: Cannot write to Excel file"
    except Exception as e:
        
        return "❌ Error: An unexpected error occurred"

if __name__ == "__main__":
    print("Enter values: Name,Age,City,Phone_number")
    user_input = input("Your values: ").strip()
    
    if user_input:
        try:
            values = [v.strip() for v in user_input.split(',')]
            if len(values) == 4:
                name, age, city, phone = values
                res = update_excel(name, int(age), city, int(phone))
                print(res)
            else:
                print("❌ Please provide exactly 4 comma-separated values")
        except:
            print("❌ Invalid format! Use: Name,Age,City,Phone_number")
    else:
        print("❌ No input provided")